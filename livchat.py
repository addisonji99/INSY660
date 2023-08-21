#Project_start_time: 2023/08/09
#Author: Krishan, Oyundari, Addison
#Course: McGill INSYS 660 Chatbot
#Client:Livmore
import pandas as pd
#Import a datepicker for the calendar GUI
import datetime
# Importing libraries for sentiment analysis
from textblob import TextBlob
from openpyxl import load_workbook
import time
import tkinter as tk
from tkinter import filedialog

#Define all the functions within the class
class LivmoreApartments:
    def __init__(self, main_menu_func):
        try:
            # If self.excel_path is not defined yet, prompt the user for the path
            if not hasattr(self, 'excel_path'):
                root = tk.Tk()
                root.withdraw() # Hide the main window
                self.excel_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
                
                if not self.excel_path: # If the user cancels the file selection, file_path will be an empty string
                    print("File selection canceled. Exiting the program.")
                    exit()
        except:
            print("File selection canceled. Exiting the program.")

        try:
            self.apartments_df = pd.read_excel(self.excel_path, sheet_name='Apartments')
            self.appointments_df = pd.read_excel(self.excel_path, sheet_name='Appointments')
            self.customers_df = pd.read_excel(self.excel_path, sheet_name='Customers')
            self.payments_df = pd.read_excel(self.excel_path, sheet_name='Payments')
            self.requests_df = pd.read_excel(self.excel_path, sheet_name='Requests')
            self.feedback_df = pd.read_excel(self.excel_path, sheet_name='Feedback')
            self.current_user_id = None
            print("Data load successfully, now you can use the bot")
        except:
            print("An error occurred while loading the Excel file. Please select the file manually.")
            
        self.main_menu_func = main_menu_func
            
    def show_apartment_details(self):
        # Getting user input for number of bedrooms and bathrooms
        print("Before book an appointment, let me understand your preference!")
        num_bedrooms = int(input("How many bedrooms would you like? "))
        num_bathrooms = int(input("How many bathrooms would you like? "))
        
        # Filtering apartments based on user input and check the availability
        filtered_apartments = self.apartments_df[(self.apartments_df['bedrooms'] == num_bedrooms) \
                            & (self.apartments_df['bathrooms'] == num_bathrooms) & (self.apartments_df['current_stock'] > 0)]
        
        # Checking if any apartments match the criteria
        if filtered_apartments.empty:
            print("Sorry, we don't have any apartments that match your criteria.")
            return
        
        while True:  
            # Outer loop to repeat the entire selection process
            # Showing available options
            print("Here are the available apartments that meet your needs:")
            print(filtered_apartments[['floor_plan_index', 'floor_plan', 'price', 'sqft']].to_string(index=False))
        
            # Inner loop to ask user to select a floor plan by index, and allow user to reselect and go to outer loop
            while True:
                floor_plan_index = int(input("Please select a floor plan by index number: "))
                selected_apartment = filtered_apartments[filtered_apartments['floor_plan_index'] == floor_plan_index]

                if not selected_apartment.empty:
                    stock = selected_apartment['current_stock'].iloc[0]
                    # Check if stock is less than 10, if yes, pop out the message say "running fast"
                    if stock < 10:
                        print(f"Nice pick! The suite is running fast, only {stock} left. Let me bring you to a short tour.")
                    else:
                        print("Nice pick! Let me bring you to a short tour.")
                    print("Virtual Tour Link:", selected_apartment['virtual_tour'].iloc[0])
                    
                    # Move to the appointment booking or reselect
                    next_action = input("What would you like to do next?\n1. I want to see other options\n2. Make an appointment\nPlease choose an option: ")
                    if next_action == '2':
                        print("Great! You can make an appointment here.") 
                        # Go to book appointment
                        self.manage_appointments('book')
                        return
                    elif next_action == '1':
                        print("Sure! Let's take a look at other options.")
                        break  # Exit the inner loop, starting the selection process over again
                    else:
                        print("Sorry I couldn't find the one you selected. Please try again.")
                        
    def book_appointment(self):
        # The appointment can be made start from next week Monday to Saturday
        upcoming_dates = [datetime.date.today() + datetime.timedelta(days=i) for i in range(7)]
        upcoming_dates = [date for date in upcoming_dates if date.weekday() < 6]
        print("Available dates for appointment:")
        for i, date in enumerate(upcoming_dates, 1):
            print(f"{i}. {date.strftime('%A, %Y-%m-%d')}")

        # User selects a date
        appointment_date = None
        while appointment_date is None:
            choice = int(input("Please choose a date by number: "))
            if 1 <= choice <= len(upcoming_dates):
                appointment_date = upcoming_dates[choice - 1].strftime('%Y-%m-%d')
            else:
                print("Invalid choice. Please try again.")

        # List of time slots
        all_time_slots = ["10:00 AM", "11:00 AM", "2:00 PM", "3:00 PM", "4:00 PM"]

        # Filtering appointments that are already booked on the selected date and have status 'Booked'
        booked_appointments = self.appointments_df[(self.appointments_df['appointment_date'] == appointment_date) & (self.appointments_df['status'] == 'Booked')]
        booked_times = booked_appointments['appointment_time'].tolist()

        # Finding available time slots by excluding booked times
        available_time_slots = [atime for atime in all_time_slots if atime not in booked_times]
        print("Available time slots:")
        for i, atime in enumerate(available_time_slots, 1):
            print(f"{i}. {atime}")

        # User selects a time slot
        selected_time = None
        while selected_time is None:
            try:
                choice = int(input("Please choose a time slot by number: "))
                if 1 <= choice <= len(available_time_slots):
                    selected_time = available_time_slots[choice - 1]
                else:
                    print("Invalid choice. Please try again.")
            except:
                print("Invalid choice. Please try again.")
                
        # Append the new booking information to the "Appointments" tab in the Excel
        new_index = len(self.appointments_df) + 1
        new_appointment = {
            'appointment_index':new_index,
            'customer_id': self.current_user_id,
            'appointment_date': appointment_date,
            'appointment_time': selected_time,
            'status': 'Booked'
        }
        
        self.appointments_df = self.appointments_df.append(new_appointment, ignore_index=True)
        
        # Open the existing workbook
        book = load_workbook(self.excel_path)

        # Create an Excel writer with the existing workbook
        writer = pd.ExcelWriter(self.excel_path, engine='openpyxl')
        writer.book = book

        # Write the updated 'Appointments' DataFrame to the existing 'Appointments' sheet
        self.appointments_df.to_excel(writer, sheet_name='Appointments', index=False)

        # Save the changes
        writer.save()
        print(f"Dear Customer,\n\nYour appointment has been booked for {appointment_date} at {selected_time}.\n\nThank you for choosing Le Livmore!")
        time.sleep(2)
        # Return to the main menu
        self.main_menu_func(self)
        return

    def manage_appointments(self, action): # action can be 'book' or 'cancel'
        if not self.current_user_id:
            self.login_customer()
        if self.current_user_id:
            if action == 'book':
                self.book_appointment()
            elif action == 'cancel':
                # Retrieve all "Booked" appointments for the logged-in customer
                booked_appointments = self.appointments_df[(self.appointments_df['customer_id'] == self.current_user_id) & (self.appointments_df['status'] == 'Booked')]

                # Check if there are any booked appointments
                if booked_appointments.empty:
                    print("You have no booked appointments to cancel.")
                    choice = input("Would you like to book an appointment? (yes/no): ")
                    if choice.lower() == 'yes':
                        self.book_appointment()
                    else:
                        print("Returning to the main menu.")
                        return
                    
                # Display the booked appointments
                print("Your booked appointments:")
                print(booked_appointments[['appointment_index', 'appointment_date', 'appointment_time']].to_string(index=False))

                # Ask the customer to choose an appointment to cancel
                while True:
                    try:
                        appointment_index = int(input("Please enter the appointment index to cancel: "))
                        selected_appointment = booked_appointments[booked_appointments['appointment_index'] == appointment_index]
                    except:
                        print("Invalid appointment index. Please try again.")
                        continue
                    if not selected_appointment.empty:
                        # Update the status to "Canceled"
                        self.appointments_df.loc[self.appointments_df['appointment_index'] == appointment_index, 'status'] = 'Canceled'
                        print("Appointment canceled successfully!")
                        break
                    else:
                        print("Invalid appointment index. Please try again.")
            else:
                print("Invalid action!")
        else:
            print("Please log in to manage appointments!")
            self.login_customer()

    def login_customer(self):
        customer = pd.DataFrame()  # Initialize customer as an empty DataFrame
        while True:
            has_account = input("Do you have an account with us? (yes/no): ")

            if has_account.lower() == 'no':
                print("Let's create an account for you!")
                # Call the apply_customer function to create a new account
                customer_id = self.apply_customer()
                self.current_user_id = customer_id
                print("Login successful!")
                #self.manage_appointments('book')  # Directly proceed to book an appointment
                return 
            elif has_account.lower() == 'yes':
                customer_id = int(input("Please login by entering your customer ID: "))
                password = input("Please enter your password: ")
                customer = self.customers_df[(self.customers_df['customer_id'] == customer_id) & (self.customers_df['password'] == password)]
                if not customer.empty:
                    self.current_user_id = customer_id
                    print("Login successful!")
                    #self.manage_appointments()  # Directly proceed to manage an appointment
                    return
                else:
                    print("We didn't find your account.")
                    choice = input("Would you like to:\n1. Try again\n2. Exit\n3. Create a new account\nPlease choose an option: ")

                    if choice == '1':
                        continue  # Continue the loop to try again
                    elif choice == '2':
                        break  # Exit the loop without logging in
                    elif choice == '3':
                        self.apply_customer()  # Call the apply_customer method to create a new account
                    else:
                        print("Invalid choice. Please try again.")
            else:
                print("Invalid option. Please try again.")
                continue  # Continue to the next iteration of the loop

                
    #Create new account
    def apply_customer(self):
        name = input("Please enter your name: ")
        email = input("Please enter your email: ")
        phone = input("Please enter your phone number: ")
        password = input("Please create a new password: ")
        customer_id = len(self.customers_df) + 1
        new_customer = {
            'name': name,
            'email': email,
            'phone': phone,
            'customer_id': customer_id,
            'password': password
        }

        # Append the new customer to the DataFrame (in memory)
        self.customers_df = self.customers_df.append(new_customer, ignore_index=True)

        # Open the Excel file with the existing workbook
        with pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a') as writer:
            writer.book = load_workbook(self.excel_path)

            # Delete the existing 'Customers' sheet
            writer.book.remove(writer.book['Customers'])

            # Write the updated customers DataFrame back to the 'Customers' sheet
            self.customers_df.to_excel(writer, sheet_name='Customers', index=False)
            time.sleep(2)
        print(f"Application successful! Your customer ID is {customer_id}. Please keep it safe for future reference.") 
        return customer_id
    
    #Create function to login as tenant
    def login_as_tenant(self):
        # Load the customer data from the Excel file
        customers_df = pd.read_excel(self.excel_path, sheet_name='Customers')

        while True:
            customer_id = input("Please enter your customer ID: ")
            password = input("Please enter your password: ")

            # Validate the credentials by matching them against the customer DataFrame
            customer = customers_df[
                (customers_df['customer_id'] == int(customer_id)) &
                (customers_df['password'] == password)
            ]

            if not customer.empty:
                print("Login successful!")
                self.current_user_id = int(customer_id)  # Set the current_user_id attribute
                # Go back to the tenant menu function directly
                time.sleep(2)
                break
            else:
                print("Login failed. Invalid customer ID or password.")
                retry_choice = input("Would you like to try again or return to the main menu? (retry/main): ")
                if retry_choice.lower() == 'main':
                    main_menu()
                    break

    def manage_payments(self):
        if self.current_user_id:
            payment_status = self.payments_df[self.payments_df['customer_id'] == self.current_user_id]

            if not payment_status.empty:
                print("Your payment detail is:")
                print(payment_status)
                print("Thank you for checking")
            else:
                print("Sorry, there is currently no payment detail.")
        else:
            print("Please log in to view payment status!")

    def manage_requests(self):
        if self.current_user_id:
            request_text = input("Please enter your request: ")
            self.requests_df = self.requests_df.append({
                'customer_id': self.current_user_id,
                'request_text': request_text,
                'request_status': 'Pending'
            }, ignore_index=True)
            print("Your request has been submitted and will be resolved in the next 24 hours.")
        else:
            print("Please log in to submit a request!")

    def manage_feedback(self):
        if self.current_user_id:
            feedback_text = input("Please provide your feedback: ")
            sentiment_analysis = TextBlob(feedback_text)
            sentiment = 'Positive' if sentiment_analysis.sentiment.polarity > 0 else ('Neutral' if sentiment_analysis.sentiment.polarity == 0 else 'Negative')
            self.feedback_df = self.feedback_df.append({
                'customer_id': self.current_user_id,
                'feedback_text': feedback_text,
                'sentiment': sentiment
            }, ignore_index=True)
            print(f"Thank you for your feedback! Your sentiment is {sentiment}.")
            time.sleep(1)
        else:
            print("Please log in to provide feedback!")
    
    def show_neighborhood(self):
        print("There are many places in the neighborhood: a, b, c, d, e, f")

    def logout(self):
        self.current_user_id = None
        print("Logged out successfully!")
        time.sleep(2)
        
    
    # Saving the updates to the Excel file
    def save_to_excel(self):
        with pd.ExcelWriter(self.excel_path, mode='w') as writer:
            self.apartments_df.to_excel(writer, sheet_name='Apartments', index=False)
            self.appointments_df.to_excel(writer, sheet_name='Appointments', index=False)
            self.customers_df.to_excel(writer, sheet_name='Customers', index=False)
            self.payments_df.to_excel(writer, sheet_name='Payments', index=False)
            self.requests_df.to_excel(writer, sheet_name='Requests', index=False)
            self.feedback_df.to_excel(writer, sheet_name='Feedback', index=False)
        print("Data saved to Excel file.")

#Initiate the bot with Menus
def about_section():
    print("\nAbout Le Livmore:")
    print("Le Livmore is all about delivering more. We are focused on offering residents more of what inspires them.")
    print("From amenities to community atmosphere and overall lifestyle, live more of your life with Le Livmore.")
    print("\nContact Information:")
    print("Phone number: 514-405-6284")
    print("Email: lelivmore@gwlra.com")
    print("\nLegal Terms and Conditions:")
    print("Visit this link for legal terms and conditions: https://www.gwlraresidential.com/privacy\n")

def help_section():
    print("\nFAQs for Livmore Apartment Customers:")
    print("Question 1: What is unique about Le Livmore?\nAnswer 1: Le Livmore is truly unique.")
    print("Question 2: What appliances and amenities are included?\nAnswer 2: Included are six appliances: refrigerator, stove, dishwasher, microwave, washer, and dryer. Hot water expenses are covered. The amenities offered encompass a pool, sauna, gym, coworking space, wine cellar, and more, in addition to the six appliances provided and covered hot water costs.")
    print("Question 3: Are pets allowed?\nAnswer 3: Yes, you feel at home, so should your pooch! Le Livmore is pet-welcoming, with a dedicated outdooe space for your pets to get all the exercise they need. There’s also a doggy spa so your pup can be completely pampered.")
    print("Question 4: Am I committing to anything if I schedule a viewing with Le Livmore?\nAnswer 4: Strictly no obligation. We are here to provide recommendations for rooms that meet your requirements.")
    print("Question 5: What is available in the neighborhood?\nAnswer 5: The Le Livmore apartment boasts a vibrant neighborhood, featuring local gems like Ferreria Café, Wienstein & Gavino’s, Biiru, Bar Le Mal Nécessaire, Centre Eaton de Montréal, the Montréal Museum of Fine Arts, and right at your doorstep, the lively summer festival hub of Place des arts.\n")

def initialize_livmore(main_menu_func):
    return LivmoreApartments(main_menu_func)

def main_menu(livmore):
    while True:
            print("\nWelcome to Le Livmore Apartments! How can I assist you with, please put the number in the box")
            print("1. Not a Tenant, Yet")
            print("2. Login as Tenant")
            print("3. Term and Conditions")
            print("4. Help")
            print("5. Exit")
            choice = input("Please choose an option: ")
            if choice == '1': # Future tenant
                user_state = 'future_tenant'
                future_tenant_menu(livmore)
            elif choice == '2': # Existing User
                user_state = 'tenant'
                livmore.login_as_tenant()
                tenant_menu(livmore)
            elif choice == '3': # Information
                about_section()
            elif choice == '4': # Help
                help_section()
            elif choice == '5': # Exit
                livmore.save_to_excel()
                print("Thank you for visiting Le Livmore Apartments! Goodbye.")
            else:
                print("Sorry I don't understand, please try again.")
            
def future_tenant_menu(livmore):
        print("\nWelcome, thank you for choosing Livmore, tell me what you are looking:")
        print("1. Book a Room Tour")
        print("2. Manage Your Bookings")
        print("3. Learn about Neighborhood")
        print("4. Exit")
        new_user_choice = input("Please choose an option: ")
        if new_user_choice == '1':
            livmore.show_apartment_details()
        elif new_user_choice == '2':
            action = input("Do you want to 'book', or 'cancel' an appointment? Shall you want to reschedule, please cancel it first and book it again")
            livmore.manage_appointments(action.lower())
        elif new_user_choice == '3':
            livmore.show_neighborhood()
        elif new_user_choice == '4':
            print("Exit successfully")
            return # Return to main menu
        else:
            print("Sorry I don't understand, please try again")
            future_tenant_menu(livmore)


def tenant_menu(livmore):
        while True:
            print("\nExisting User Options:")
            print("1. Payment")
            print("2. Request")
            print("3. Feedback")
            print("4. Logout")
            existing_user_choice = input("Please choose an option: ")

            if existing_user_choice == '1':
                livmore.manage_payments()
            elif existing_user_choice == '2':
                livmore.manage_requests()
            elif existing_user_choice == '3':
                livmore.manage_feedback()
            elif existing_user_choice == '4':
                livmore.logout()
                return # Return to main menu
            else:
                print("Sorry I don't understand, please try again")

# Run the main menu
livmore = initialize_livmore(main_menu)
main_menu(livmore)

